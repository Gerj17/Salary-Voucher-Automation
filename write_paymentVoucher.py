import datetime
from pandas.tseries.offsets import BMonthEnd
from read_wageSheet import information, month
from excel_to_pdf import ExcelPrintPdf
from pathlib import Path as p
import time
import openpyxl.utils.cell as opx
import openpyxl
import project_inputs

#  Code to get the last day of the business week
d = datetime.date.today()
offset = BMonthEnd()
last_bs_day = offset.rollforward(d).strftime("%d")
theYear = offset.rollforward(d).strftime("%Y")

loc_excel_docs = p.cwd()  # Create a path to directory with documents
name = r'Ideal Bakery Wage Payment Voucher.xlsx'
wb_pay = openpyxl.load_workbook(
    p.joinpath(loc_excel_docs, name), )  # Open excel

ws = wb_pay['Main']

date = 'c8'
payee = 'c14'
description = 'c17'
amount = 'c22'
value = 'k25'
voucher_No = 'k8'
cheque_No = 'k12'
v_num = project_inputs.file_voucher_number
c_num = project_inputs.file_cheque_number

# splitter = lambda x :opx.coordinate_from_string(x)


# new = ws['h8'] = Test_Text
# go = ws['h8'].value

# Alston Prince ('Alston Prince', 1906.21, None, datetime.datetime(2021, 5, 31, 0, 0))
# Carl Smith ('Carl Smith', 1464.14, 664.1400000000001, datetime.datetime(2021, 5, 31, 0, 0))

counter = 0
for key, value_ in information.items():
    print(information[key], len(information[key]))
    ws[payee] = (information[key][0])  # Name of emplyee
    if information[key][2] is not None:
        ws[value] = (information[key][2])  # fortnight pay
        ws[description] = f"2nd Fortnight Pay for Month the of {month} 2021"  # Description
        print('is integer')
    else:
        # print(type(information[key][1])) # Full Salary
        ws[value] = (information[key][1])  # Full Salary
        ws[description] = f"Pay for Month the of {month} 2021"  # Description
        print(" No Fortnight pay")
        pass
    ws[amount] = '---'
    ws[date] = information[key][3]

    ws[voucher_No] = v_num + counter
    ws[cheque_No] = c_num + counter
    counter += 1

    try:
        wb_pay.save(name)
        print('Document saved')
    except PermissionError:
        print(f"\nFile >{name}< is probably open\n Close file and try again \n ")
        time.sleep(5)

    # ExcelPrintPdf(loc_excel_docs.joinpath(name), key, )
    # ExcelPrintPdf(loc_excel_docs.joinpath(name), f'Ideal Bakery Payment Voucher {v_num + counter}-', )
    ExcelPrintPdf(loc_excel_docs.joinpath(name),
                  f'Ideal Bakery Payment Voucher {v_num + counter}-{last_bs_day} {month.capitalize()} {theYear}', )

    print('document converted \n ')
    time.sleep(1)

#testing = information["Carl Smith"][3]
# print(testing, type(testing))
# st_date = testing.strtime("%d %B %Y")

#example = 'Ideal Bakery Payment Voucher 201-30 May 2021'
#print(f'Ideal Bakery Payment Voucher {v_num + counter}-{last_bs_day} {month.capitalize()} {theYear}')
#print(example)
# TODO Create the loop to do all employees
# TODO Create the logic for voucher and cheque numbers
# TODO Create the logic for the save name of the pdf files
# TODO Maybe create GUI for input data or maybe just a file to take the relative data and use time.sleep() to shot
# TODO
# TODO
