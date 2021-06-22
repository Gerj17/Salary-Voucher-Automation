import datetime
import project_inputs
from pathlib import Path as p
import openpyxl
from openpyxl import utils
import time

date = str(project_inputs.file_date).strip()  # format in number (ddmmyy) all numbers
month = project_inputs.file_month

loc_excel_docs = p.cwd().joinpath()  # Create a path to directory with documents
wb_wage = openpyxl.load_workbook(
    p.joinpath(loc_excel_docs, f'{date} IDEAL BAKERY _PAYROLL {month.upper()}.xlsx'), data_only=True)  # Open excel
ws = wb_wage['Pay slips']

# TODO: create loop to get each employees data

name_1 = ws["B5"].value  # payee
date_1 = ws["C5"].value  # TODO fix with daytime
cheque_num = None
voucher_nun = None
description = None  # pay monthly or fortnight only end of month salaries do here
amount = '---'  # place holder for no value in sheet to be printed
net_pay_1 = ws["D22"].value  # end of month pay
fortnight_pay = ws["F22"].value

# counter = 24
st_name_1 = ["B5"]
st_date_1 = ["C5"]
st_net_pay_1 = ["D22"]
st_fortnight_pay = ["F22"]
# print(spl_a(st_name_1)) Testing fuction
# print(spl_b(st_name_1)) Testion Fuction
spl_a = lambda value: value[0][:1]  # get number from variable "b5" = b
spl_b = lambda value: value[0][1:]  # get letter from variable "b5" = 5

finisher = []


def runner(item, counter=24):
    """

    :param item: list containing only Excel Cell Address
    :param counter:
    :return:
    """
    # print('item input', item)
    # print('testing', ws[st_name_1[0]].value)
    Sp_sum = (int(spl_b(item)) + counter)  # sums row numbers after splitting out column letter
    new_cell = str(f'{spl_a(item)}' + str(Sp_sum))  # add new row number and column number together
    conCat = ws[new_cell].value  # add new row number and column number together
    # print('new cell path ', new_cell,'\n')
    finisher.append(conCat)
    item.clear()
    item.append(new_cell)
    # print("Cel value->", conCat,'------',new_cell)
    # print("Cel value->", conCat, )
    # print("Cel value", conCat,'------',type(conCat) )
    # print(conCat is None)
    return conCat


var = "#REF!"  # Value from excel document

information = {}

while True:
    # for x in range(1):
    try:
        # print('tester :', finisher[-1])
        if int is type(finisher[-1]):
            # print("INTEGER!!!!")
            # break
            pass
    except IndexError:
        print("IndexError")
    finally:
        pass
    if var in finisher:
        break
    else:
        name = runner(st_name_1)
        NetPay = runner(st_net_pay_1)
        FortNightPay = runner(st_fortnight_pay)
        Date = runner(st_date_1)

        information[name] = name, NetPay, FortNightPay, Date  # add personal info to dictionary        print('\n')
        # print('---', finisher)
        # print('\n')
    # if

information[name_1] = name_1, net_pay_1, fortnight_pay, date_1

information.pop("#REF!")

